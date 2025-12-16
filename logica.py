"""
Lógica de Negocio del Sistema de Bazar - VERSIÓN LIMPIA
Manejo de CSV, validaciones y cálculos
"""
import csv
import os
from datetime import datetime
from config import *
import json 

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    OPENPYXL_DISPONIBLE = True
except ImportError:
    OPENPYXL_DISPONIBLE = False
    print("⚠️ Advertencia: openpyxl no está instalado. No se generarán archivos Excel.")
    print("   Para instalar: pip install openpyxl")


class GestorProductos:
    """Maneja la carga, guardado y manipulación de productos"""
    
    def __init__(self):
        self.productos = []
        self.cargar_productos()
    
    def cargar_productos(self):
        """Carga los productos desde el CSV"""
        if not os.path.exists(RUTA_PRODUCTOS):
            self.crear_csv_ejemplo()
        
        try:
            with open(RUTA_PRODUCTOS, 'r', encoding='utf-8') as archivo:
                lector = csv.DictReader(archivo)
                self.productos = list(lector)
                # Convertir precios y stock a numéricos
                for producto in self.productos:
                    producto['precio'] = float(producto['precio'])
                    # Si no existe la columna stock, agregarla con valor 0
                    if 'stock' not in producto:
                        producto['stock'] = 0
                    else:
                        try:
                            producto['stock'] = int(producto['stock'])
                        except:
                            producto['stock'] = 0
            return True
        except Exception as e:
            print(f"Error al cargar productos: {e}")
            return False
    
    def crear_csv_ejemplo(self):
        """Crea un CSV de ejemplo con productos iniciales"""
        productos_ejemplo = [
            {'codigo': '001', 'nombre': 'Coca Cola 500ml', 'precio': '3.50', 'categoria': 'Bebidas', 'stock': '50'},
            {'codigo': '002', 'nombre': 'Inca Kola 500ml', 'precio': '3.50', 'categoria': 'Bebidas', 'stock': '40'},
            {'codigo': '003', 'nombre': 'Agua San Luis 625ml', 'precio': '2.00', 'categoria': 'Bebidas', 'stock': '60'},
            {'codigo': '004', 'nombre': 'Galletas Oreo', 'precio': '4.50', 'categoria': 'Snacks', 'stock': '30'},
            {'codigo': '005', 'nombre': 'Papas Lays', 'precio': '5.00', 'categoria': 'Snacks', 'stock': '25'},
            {'codigo': '006', 'nombre': 'Chocolate Sublime', 'precio': '2.50', 'categoria': 'Dulces', 'stock': '45'},
            {'codigo': '007', 'nombre': 'Chicles Trident', 'precio': '1.50', 'categoria': 'Dulces', 'stock': '100'},
            {'codigo': '008', 'nombre': 'Pan Integral', 'precio': '6.00', 'categoria': 'Panadería', 'stock': '15'},
        ]
        
        with open(RUTA_PRODUCTOS, 'w', newline='', encoding='utf-8') as archivo:
            escritor = csv.DictWriter(archivo, fieldnames=COLUMNAS_PRODUCTOS)
            escritor.writeheader()
            escritor.writerows(productos_ejemplo)
    
    def guardar_productos(self):
        """Guarda los productos en el CSV"""
        try:
            with open(RUTA_PRODUCTOS, 'w', newline='', encoding='utf-8') as archivo:
                escritor = csv.DictWriter(archivo, fieldnames=COLUMNAS_PRODUCTOS)
                escritor.writeheader()
                escritor.writerows(self.productos)
            return True
        except Exception as e:
            print(f"Error al guardar productos: {e}")
            return False
    
    def agregar_producto(self, codigo, nombre, precio, categoria, stock=0):
        """Agrega un nuevo producto"""
        # Verificar si el código ya existe
        if any(p['codigo'] == codigo for p in self.productos):
            return False, "El código ya existe"
        
        nuevo_producto = {
            'codigo': codigo,
            'nombre': nombre,
            'precio': float(precio),
            'categoria': categoria,
            'stock': int(stock)
        }
        self.productos.append(nuevo_producto)
        return self.guardar_productos(), "Producto agregado exitosamente"
    
    def editar_producto(self, codigo, nombre, precio, categoria, stock=None):
        """Edita un producto existente"""
        for producto in self.productos:
            if producto['codigo'] == codigo:
                producto['nombre'] = nombre
                producto['precio'] = float(precio)
                producto['categoria'] = categoria
                if stock is not None:
                    producto['stock'] = int(stock)
                return self.guardar_productos(), "Producto editado exitosamente"
        return False, "Producto no encontrado"
    
    def eliminar_producto(self, codigo):
        """Elimina un producto"""
        self.productos = [p for p in self.productos if p['codigo'] != codigo]
        return self.guardar_productos()
    
    def buscar_producto(self, codigo):
        """Busca un producto por código"""
        for producto in self.productos:
            if producto['codigo'] == codigo:
                return producto
        return None
    
    def obtener_nombres_productos(self):
        """Retorna lista de nombres con precio para el combobox"""
        resultado = []
        for p in self.productos:
            texto = f"{p['nombre']} - S/ {p['precio']:.2f}"
            # Si el stock está activado, mostrar stock
            if STOCK_ACTIVADO:
                texto += f" [STOCK: {p['stock']}]"
            resultado.append(texto)
        return resultado
    
    def obtener_categorias(self):
        """Retorna lista de categorías únicas"""
        categorias = set(p['categoria'] for p in self.productos)
        return sorted(list(categorias))
    
    def actualizar_stock(self, codigo, cantidad_vendida):
        """Actualiza el stock de un producto después de una venta"""
        if not STOCK_ACTIVADO:
            return True, "Stock desactivado"
        
        for producto in self.productos:
            if producto['codigo'] == codigo:
                nuevo_stock = producto['stock'] - cantidad_vendida
                if nuevo_stock < 0:
                    return False, f"Stock insuficiente. Disponible: {producto['stock']}"
                producto['stock'] = nuevo_stock
                self.guardar_productos()
                
                # Advertencia si el stock es bajo
                if nuevo_stock <= 5:
                    return True, f"ADVERTENCIA: Stock bajo para '{producto['nombre']}': {nuevo_stock} unidades"
                return True, "Stock actualizado"
        return False, "Producto no encontrado"
    
    def productos_stock_bajo(self, umbral=10):
        """Retorna productos con stock bajo"""
        if not STOCK_ACTIVADO:
            return []
        
        return [p for p in self.productos if p['stock'] <= umbral]
    
    def obtener_siguiente_codigo_variable(self):
        """
        Genera el siguiente código para productos variables (VAR001, VAR002, etc.)
        """
        codigos_var = [p['codigo'] for p in self.productos if p['codigo'].startswith('VAR')]
        
        if not codigos_var:
            return 'VAR001'
        
        # Extraer números y encontrar el máximo
        numeros = []
        for cod in codigos_var:
            try:
                num = int(cod.replace('VAR', ''))
                numeros.append(num)
            except:
                continue
        
        if not numeros:
            return 'VAR001'
        
        siguiente = max(numeros) + 1
        return f'VAR{siguiente:03d}'  # VAR001, VAR002, etc.


class GestorVentas:
    """Maneja el registro de ventas y cálculos"""
    
    def __init__(self):
        self.ventas_actuales = []
    
    def agregar_venta(self, producto, cantidad, metodo_pago):
        """Agrega una venta a la lista actual"""
        venta = {
            'codigo': producto['codigo'],
            'nombre': producto['nombre'],
            'cantidad': cantidad,
            'precio_unitario': producto['precio'],
            'subtotal': producto['precio'] * cantidad,
            'metodo_pago': metodo_pago,
            'categoria': producto['categoria']
        }
        self.ventas_actuales.append(venta)
        return venta
    def autoguardar_temporal(self):
        """Guarda automáticamente las ventas en archivo temporal (en segundo plano)"""
        if not self.ventas_actuales:
            return
        
        try:
            # Calcular totales para el resumen
            totales = self.calcular_totales()
            
            datos_temporal = {
                'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'total': totales['general'],
                'cantidad_productos': len(self.ventas_actuales),
                'ventas': self.ventas_actuales
            }
            
            with open(RUTA_TEMPORAL, 'w', encoding='utf-8') as f:
                json.dump(datos_temporal, f, ensure_ascii=False, indent=2)
        except Exception as e:
            # Silencioso - no interrumpir flujo normal si falla
            print(f"Error en autoguardado temporal: {e}") 


    def cargar_temporal(self):
        """
        Carga ventas desde archivo temporal si existe
        Returns: (exito: bool, datos: dict o None)
        """
        if not os.path.exists(RUTA_TEMPORAL):
            return False, None
        
        try:
            with open(RUTA_TEMPORAL, 'r', encoding='utf-8') as f:
                datos = json.load(f)
            
            # Verificar que tenga ventas
            if not datos.get('ventas'):
                self.limpiar_temporal()
                return False, None
            
            # Cargar ventas (sin modificar stock)
            self.ventas_actuales = datos['ventas']
            
            return True, datos
        except Exception as e:
            print(f"Error al cargar temporal: {e}")
            return False, None



    def limpiar_temporal(self):
        """Elimina el archivo temporal"""
        try:
            if os.path.exists(RUTA_TEMPORAL):
                os.remove(RUTA_TEMPORAL)
        except Exception as e:
            print(f"Error al limpiar temporal: {e}")



    def eliminar_venta(self, indice):
        """Elimina una venta de la lista actual"""
        if 0 <= indice < len(self.ventas_actuales):
            self.ventas_actuales.pop(indice)
            return True
        return False
    
    def calcular_totales(self):
        """Calcula los totales de las ventas actuales"""
        total_general = sum(v['subtotal'] for v in self.ventas_actuales)
        total_efectivo = sum(v['subtotal'] for v in self.ventas_actuales if v['metodo_pago'] == 'E')
        total_virtual = sum(v['subtotal'] for v in self.ventas_actuales 
                           if v['metodo_pago'] in METODOS_VIRTUALES)
        
        return {
            'general': total_general,
            'efectivo': total_efectivo,
            'virtual': total_virtual
        }
    
    def guardar_ventas(self):
        """Guarda las ventas EN CSV (Excel_app) y EXCEL (Excel_registro)"""
        if not self.ventas_actuales:
            return False, "No hay ventas para guardar"
        
        fecha_actual = datetime.now()
        fecha_str = fecha_actual.strftime('%Y-%m-%d')
        hora_str = fecha_actual.strftime('%H:%M:%S')
        
        # 1. GUARDAR CSV EN Excel_app (para el programa)
        nombre_csv = os.path.join(RUTA_VENTAS_APP, f'ventas_{fecha_str}.csv')
        archivo_existe = os.path.exists(nombre_csv)
        
        try:
            with open(nombre_csv, 'a', newline='', encoding='utf-8') as archivo:
                escritor = csv.DictWriter(archivo, fieldnames=COLUMNAS_VENTAS)
                
                if not archivo_existe:
                    escritor.writeheader()
                
                # Agregar fecha y hora a cada venta
                for venta in self.ventas_actuales:
                    venta['fecha'] = fecha_str
                    venta['hora'] = hora_str
                    escritor.writerow(venta)
        except Exception as e:
            return False, f"Error al guardar CSV: {e}"
        
        # 2. GENERAR EXCEL EN Excel_registro (para visualización)
        if OPENPYXL_DISPONIBLE:
            try:
                self._generar_excel_visual(fecha_str)
            except Exception as e:
                print(f"⚠️ Error al generar Excel visual: {e}")
        
        # Calcular totales antes de limpiar
        totales = self.calcular_totales()
        num_productos = len(self.ventas_actuales)
        
        # Limpiar lista automáticamente después de guardar
        self.ventas_actuales = []
        
        return True, {
            'archivo': nombre_csv,
            'total': totales['general'],
            'efectivo': totales['efectivo'],
            'virtual': totales['virtual'],
            'productos': num_productos
        }
    
    def _generar_excel_visual(self, fecha_str):
        """Genera un Excel visual para registro humano"""
        nombre_excel = os.path.join(RUTA_VENTAS_REGISTRO, f'ventas_{fecha_str}.xlsx')
        
        # Agrupar ventas por método de pago
        ventas_efectivo = []
        ventas_digitales = []
        
        for venta in self.ventas_actuales:
            if venta['metodo_pago'] == 'E':
                # Buscar si ya existe este producto en efectivo
                encontrado = False
                for v in ventas_efectivo:
                    if v['nombre'] == venta['nombre']:
                        v['cantidad'] += venta['cantidad']
                        v['subtotal'] += venta['subtotal']
                        encontrado = True
                        break
                if not encontrado:
                    ventas_efectivo.append({
                        'nombre': venta['nombre'],
                        'cantidad': venta['cantidad'],
                        'subtotal': venta['subtotal']
                    })
            else:
                # Método digital
                ventas_digitales.append({
                    'metodo': METODOS_PAGO.get(venta['metodo_pago'], 'Otros'),
                    'monto': venta['subtotal']
                })
        
        # Crear Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Ventas"
        
        # Anchos de columna
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 5
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 15
        
        # ENCABEZADOS
        ws['A1'] = 'Productos'
        ws['B1'] = 'Efectivo'
        ws['D1'] = ' '
        ws['E1'] = 'Yape / Plin'
        
        # Formato encabezados (opcional, simple)
        for cell in ['A1', 'B1', 'D1', 'E1']:
            ws[cell].font = Font(bold=True)
            ws[cell].alignment = Alignment(horizontal='center')
        
        # PRODUCTOS (EFECTIVO)
        fila = 2
        total_efectivo = 0
        for v in ventas_efectivo:
            ws[f'A{fila}'] = f"{v['nombre']} x{v['cantidad']}"
            ws[f'B{fila}'] = f"S/ {v['subtotal']:.2f}"
            ws[f'B{fila}'].alignment = Alignment(horizontal='right')
            total_efectivo += v['subtotal']
            fila += 1
        
        # PAGOS DIGITALES
        fila_digital = 2
        total_digital = 0
        for v in ventas_digitales:
            ws[f'E{fila_digital}'] = f"S/ {v['monto']:.2f}"
            ws[f'E{fila_digital}'].alignment = Alignment(horizontal='right')
            total_digital += v['monto']
            fila_digital += 1
        
        # TOTALES
        fila_total = max(fila, fila_digital) + 1
        ws[f'A{fila_total}'] = 'TOTAL'
        ws[f'B{fila_total}'] = f"S/ {total_efectivo:.2f}"
        ws[f'E{fila_total}'] = f"S/ {total_digital:.2f}"
        
        # Formato totales (opcional, simple)
        for cell in [f'A{fila_total}', f'B{fila_total}', f'E{fila_total}']:
            ws[cell].font = Font(bold=True)
            ws[cell].alignment = Alignment(horizontal='right' if cell != f'A{fila_total}' else 'left')
        
        # Guardar
        wb.save(nombre_excel)
        print(f"✅ Excel visual guardado en: {nombre_excel}")
    
    def limpiar_ventas(self):
        """Limpia la lista de ventas actuales (EMERGENCIA)"""
        self.ventas_actuales = []
        self.limpiar_temporal() 
    
    def obtener_ventas(self):
        """Retorna la lista de ventas actuales"""
        return self.ventas_actuales
    
    def obtener_historial(self, fecha=None):
        """Lee el historial de ventas de una fecha específica o todas"""
        historial = []
        
        if fecha:
            # Leer archivo específico
            nombre_archivo = os.path.join(RUTA_VENTAS, f'ventas_{fecha}.csv')
            if os.path.exists(nombre_archivo):
                try:
                    with open(nombre_archivo, 'r', encoding='utf-8') as archivo:
                        lector = csv.DictReader(archivo)
                        historial = list(lector)
                except Exception as e:
                    print(f"Error al leer historial: {e}")
        else:
            # Leer todos los archivos de ventas
            if os.path.exists(RUTA_VENTAS):
                for archivo in sorted(os.listdir(RUTA_VENTAS)):
                    if archivo.startswith('ventas_') and archivo.endswith('.csv'):
                        ruta_completa = os.path.join(RUTA_VENTAS, archivo)
                        try:
                            with open(ruta_completa, 'r', encoding='utf-8') as f:
                                lector = csv.DictReader(f)
                                historial.extend(list(lector))
                        except Exception as e:
                            print(f"Error al leer {archivo}: {e}")
        
        return historial


def validar_numero(valor, tipo='int'):
    """Valida que un valor sea numérico"""
    try:
        if tipo == 'int':
            return int(valor) > 0
        elif tipo == 'float':
            return float(valor) > 0
    except ValueError:
        return False
    return False